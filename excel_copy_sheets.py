from openpyxl import load_workbook, Workbook
from copy import copy

# Load the existing workbook
source_wb = load_workbook('E:/Downloads/BP_CUSTOMER Maping Sheet 1.xlsx')

# Create a new workbook
dest_wb = Workbook()

# Specify the sheets you want to copy
sheets_to_copy = ['Businesspatner', 'to_customer', 'to_CustomerTaxGrouping']

for sheet_name in sheets_to_copy:
    source_ws = source_wb[sheet_name]
    dest_ws = dest_wb.create_sheet(title=sheet_name)

    # Copy all cells and their styles
    for row in source_ws.iter_rows():
        for cell in row:
            new_cell = dest_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = copy(cell.number_format)
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)

# Remove the default sheet created by openpyxl if it's not needed
if 'Sheet' in dest_wb.sheetnames:
    del dest_wb['Sheet']

# Save the new workbook
dest_wb.save('E:/Downloads/changes needed.xlsx')
